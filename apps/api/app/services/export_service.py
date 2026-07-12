"""
Export Service for Synthesize.io
================================
Multi-format export support for generated datasets.

Supported formats:
- CSV (with configurable delimiters)
- Excel (.xlsx) with formatting
- Parquet (with compression)
- SQL (INSERT statements with batching)
- JSON (line-delimited and standard)
"""

import csv
import json
import io
import os
import logging
from typing import Dict, Any, List, Optional, Iterator, Union
from datetime import datetime, date
from decimal import Decimal
from uuid import UUID
from enum import Enum

logger = logging.getLogger(__name__)

# Export format enum
class ExportFormat(str, Enum):
    CSV = "csv"
    EXCEL = "xlsx"
    PARQUET = "parquet"
    SQL = "sql"
    JSON = "json"
    JSONL = "jsonl"  # Line-delimited JSON


class ExportOptions:
    """Configuration options for export operations."""
    
    def __init__(
        self,
        format: ExportFormat = ExportFormat.CSV,
        # CSV options
        csv_delimiter: str = ",",
        csv_quotechar: str = '"',
        csv_include_header: bool = True,
        csv_encoding: str = "utf-8",
        # Excel options
        excel_sheet_name: str = "Data",
        excel_freeze_header: bool = True,
        excel_auto_column_width: bool = True,
        # Parquet options
        parquet_compression: str = "snappy",  # snappy, gzip, brotli, none
        parquet_row_group_size: int = 100000,
        # SQL options
        sql_table_name: str = "data",
        sql_dialect: str = "postgresql",  # postgresql, mysql, sqlite
        sql_batch_size: int = 1000,
        sql_include_create: bool = True,
        sql_include_drop: bool = False,
        # JSON options
        json_indent: Optional[int] = None,  # None for compact
        json_orient: str = "records",  # records, columns, values
    ):
        self.format = format
        self.csv_delimiter = csv_delimiter
        self.csv_quotechar = csv_quotechar
        self.csv_include_header = csv_include_header
        self.csv_encoding = csv_encoding
        self.excel_sheet_name = excel_sheet_name
        self.excel_freeze_header = excel_freeze_header
        self.excel_auto_column_width = excel_auto_column_width
        self.parquet_compression = parquet_compression
        self.parquet_row_group_size = parquet_row_group_size
        self.sql_table_name = sql_table_name
        self.sql_dialect = sql_dialect
        self.sql_batch_size = sql_batch_size
        self.sql_include_create = sql_include_create
        self.sql_include_drop = sql_include_drop
        self.json_indent = json_indent
        self.json_orient = json_orient


class ExportService:
    """Service for exporting datasets to various formats."""
    
    def __init__(self, options: Optional[ExportOptions] = None):
        self.options = options or ExportOptions()
    
    def export(
        self,
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, Any]] = None,
        format: Optional[ExportFormat] = None
    ) -> bytes:
        """
        Export data to the specified format.
        
        Args:
            data: List of dictionaries containing the data
            schema: Optional schema definition for type inference
            format: Override format from options
        
        Returns:
            Bytes of the exported file
        """
        export_format = format or self.options.format
        
        if export_format == ExportFormat.CSV:
            return self.export_csv(data)
        elif export_format == ExportFormat.EXCEL:
            return self.export_excel(data, schema)
        elif export_format == ExportFormat.PARQUET:
            return self.export_parquet(data, schema)
        elif export_format == ExportFormat.SQL:
            return self.export_sql(data, schema)
        elif export_format == ExportFormat.JSON:
            return self.export_json(data)
        elif export_format == ExportFormat.JSONL:
            return self.export_jsonl(data)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
    
    def get_content_type(self, format: Optional[ExportFormat] = None) -> str:
        """Get MIME content type for format."""
        export_format = format or self.options.format
        content_types = {
            ExportFormat.CSV: "text/csv",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.PARQUET: "application/octet-stream",
            ExportFormat.SQL: "application/sql",
            ExportFormat.JSON: "application/json",
            ExportFormat.JSONL: "application/x-ndjson",
        }
        return content_types.get(export_format, "application/octet-stream")
    
    def get_file_extension(self, format: Optional[ExportFormat] = None) -> str:
        """Get file extension for format."""
        export_format = format or self.options.format
        extensions = {
            ExportFormat.CSV: "csv",
            ExportFormat.EXCEL: "xlsx",
            ExportFormat.PARQUET: "parquet",
            ExportFormat.SQL: "sql",
            ExportFormat.JSON: "json",
            ExportFormat.JSONL: "jsonl",
        }
        return extensions.get(export_format, "dat")
    
    # =========================================================================
    # CSV EXPORT
    # =========================================================================
    
    def export_csv(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to CSV format."""
        if not data:
            return b""
        
        output = io.StringIO()
        
        # Get all unique field names preserving order
        fieldnames = list(data[0].keys())
        
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=self.options.csv_delimiter,
            quotechar=self.options.csv_quotechar,
            quoting=csv.QUOTE_MINIMAL,
        )
        
        if self.options.csv_include_header:
            writer.writeheader()
        
        for row in data:
            # Convert complex types to strings
            processed_row = {}
            for key, value in row.items():
                processed_row[key] = self._serialize_value(value)
            writer.writerow(processed_row)
        
        return output.getvalue().encode(self.options.csv_encoding)
    
    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================
    
    def export_excel(
        self,
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Export data to Excel (.xlsx) format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if not data:
            # Return empty workbook
            wb = openpyxl.Workbook()
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.options.excel_sheet_name
        
        # Get field names
        fieldnames = list(data[0].keys())
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = row_data.get(field)
                processed_value = self._excel_serialize_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=processed_value)
                cell.border = thin_border
        
        # Auto-adjust column widths
        if self.options.excel_auto_column_width:
            for col_idx, field in enumerate(fieldnames, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(field))
                
                for row_idx in range(2, min(len(data) + 2, 102)):  # Sample first 100 rows
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        if self.options.excel_freeze_header:
            ws.freeze_panes = "A2"
        
        # Add autofilter
        ws.auto_filter.ref = ws.dimensions
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # =========================================================================
    # PARQUET EXPORT
    # =========================================================================
    
    def export_parquet(
        self,
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Export data to Parquet format with compression."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")
        
        if not data:
            return b""
        
        # Convert to columnar format
        columns = {}
        fieldnames = list(data[0].keys())
        
        for field in fieldnames:
            columns[field] = [self._parquet_serialize_value(row.get(field)) for row in data]
        
        # Infer PyArrow types
        pa_schema = self._infer_pyarrow_schema(columns, schema)
        
        # Create PyArrow table
        table = pa.table(columns, schema=pa_schema)
        
        # Write to buffer
        output = io.BytesIO()
        
        compression = self.options.parquet_compression
        if compression == "none":
            compression = None
        
        pq.write_table(
            table,
            output,
            compression=compression,
            row_group_size=self.options.parquet_row_group_size,
        )
        
        return output.getvalue()
    
    def _infer_pyarrow_schema(
        self,
        columns: Dict[str, List],
        schema: Optional[Dict[str, Any]] = None
    ):
        """Infer PyArrow schema from data and optional schema definition."""
        import pyarrow as pa
        
        fields = []
        
        for field_name, values in columns.items():
            # Find first non-None value to infer type
            sample_value = None
            for v in values:
                if v is not None:
                    sample_value = v
                    break
            
            # Infer type
            if sample_value is None:
                pa_type = pa.string()
            elif isinstance(sample_value, bool):
                pa_type = pa.bool_()
            elif isinstance(sample_value, int):
                pa_type = pa.int64()
            elif isinstance(sample_value, float):
                pa_type = pa.float64()
            elif isinstance(sample_value, (datetime, date)):
                pa_type = pa.timestamp('us')
            else:
                pa_type = pa.string()
            
            fields.append(pa.field(field_name, pa_type))
        
        return pa.schema(fields)
    
    # =========================================================================
    # SQL EXPORT
    # =========================================================================
    
    def export_sql(
        self,
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Export data to SQL INSERT statements."""
        if not data:
            return b""
        
        output = io.StringIO()
        fieldnames = list(data[0].keys())
        table_name = self._escape_identifier(self.options.sql_table_name)
        
        # Generate DROP TABLE if requested
        if self.options.sql_include_drop:
            output.write(f"DROP TABLE IF EXISTS {table_name};\n\n")
        
        # Generate CREATE TABLE if requested
        if self.options.sql_include_create:
            create_sql = self._generate_create_table(fieldnames, data, schema)
            output.write(create_sql)
            output.write("\n\n")
        
        # Generate INSERT statements in batches
        batch_size = self.options.sql_batch_size
        columns_str = ", ".join(self._escape_identifier(f) for f in fieldnames)
        
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            
            output.write(f"INSERT INTO {table_name} ({columns_str}) VALUES\n")
            
            value_rows = []
            for row in batch:
                values = []
                for field in fieldnames:
                    value = row.get(field)
                    values.append(self._sql_serialize_value(value))
                value_rows.append(f"  ({', '.join(values)})")
            
            output.write(",\n".join(value_rows))
            output.write(";\n\n")
        
        return output.getvalue().encode('utf-8')
    
    def _generate_create_table(
        self,
        fieldnames: List[str],
        data: List[Dict[str, Any]],
        schema: Optional[Dict[str, Any]] = None
    ) -> str:
        """Generate CREATE TABLE statement."""
        table_name = self._escape_identifier(self.options.sql_table_name)
        
        columns = []
        for field in fieldnames:
            # Find first non-None value to infer type
            sample_value = None
            for row in data[:100]:  # Sample first 100 rows
                if row.get(field) is not None:
                    sample_value = row.get(field)
                    break
            
            sql_type = self._infer_sql_type(sample_value, self.options.sql_dialect)
            escaped_field = self._escape_identifier(field)
            columns.append(f"  {escaped_field} {sql_type}")
        
        return f"CREATE TABLE {table_name} (\n" + ",\n".join(columns) + "\n);"
    
    def _infer_sql_type(self, value: Any, dialect: str = "postgresql") -> str:
        """Infer SQL column type from Python value."""
        if value is None:
            return "TEXT"
        elif isinstance(value, bool):
            return "BOOLEAN"
        elif isinstance(value, int):
            if -2147483648 <= value <= 2147483647:
                return "INTEGER"
            return "BIGINT"
        elif isinstance(value, float):
            return "DOUBLE PRECISION" if dialect == "postgresql" else "DOUBLE"
        elif isinstance(value, Decimal):
            return "DECIMAL(18, 4)"
        elif isinstance(value, datetime):
            return "TIMESTAMP" if dialect == "postgresql" else "DATETIME"
        elif isinstance(value, date):
            return "DATE"
        elif isinstance(value, (list, dict)):
            return "JSONB" if dialect == "postgresql" else "JSON"
        else:
            # Estimate varchar length
            str_len = len(str(value))
            if str_len <= 50:
                return "VARCHAR(100)"
            elif str_len <= 200:
                return "VARCHAR(500)"
            else:
                return "TEXT"
    
    def _escape_identifier(self, identifier: str) -> str:
        """Escape SQL identifier based on dialect."""
        dialect = self.options.sql_dialect
        if dialect == "mysql":
            return f"`{identifier}`"
        else:  # postgresql, sqlite
            return f'"{identifier}"'
    
    def _sql_serialize_value(self, value: Any) -> str:
        """Serialize value for SQL INSERT statement."""
        if value is None:
            return "NULL"
        elif isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        elif isinstance(value, (int, float, Decimal)):
            return str(value)
        elif isinstance(value, (datetime, date)):
            return f"'{value.isoformat()}'"
        elif isinstance(value, (list, dict)):
            json_str = json.dumps(value).replace("'", "''")
            return f"'{json_str}'"
        else:
            # Escape single quotes
            str_value = str(value).replace("'", "''")
            return f"'{str_value}'"
    
    # =========================================================================
    # JSON EXPORT
    # =========================================================================
    
    def export_json(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to JSON format."""
        if self.options.json_orient == "records":
            output = data
        elif self.options.json_orient == "columns":
            if not data:
                output = {}
            else:
                output = {key: [row.get(key) for row in data] for key in data[0].keys()}
        elif self.options.json_orient == "values":
            if not data:
                output = []
            else:
                fieldnames = list(data[0].keys())
                output = {
                    "columns": fieldnames,
                    "data": [[row.get(f) for f in fieldnames] for row in data]
                }
        else:
            output = data
        
        return json.dumps(
            output,
            indent=self.options.json_indent,
            default=self._json_serializer,
            ensure_ascii=False
        ).encode('utf-8')
    
    def export_jsonl(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to line-delimited JSON (JSONL) format."""
        lines = []
        for row in data:
            line = json.dumps(row, default=self._json_serializer, ensure_ascii=False)
            lines.append(line)
        return "\n".join(lines).encode('utf-8')
    
    # =========================================================================
    # HELPER METHODS
    # =========================================================================
    
    def _serialize_value(self, value: Any) -> str:
        """Serialize value to string for CSV."""
        if value is None:
            return ""
        elif isinstance(value, bool):
            return "true" if value else "false"
        elif isinstance(value, (datetime, date)):
            return value.isoformat()
        elif isinstance(value, (list, dict)):
            return json.dumps(value)
        elif isinstance(value, UUID):
            return str(value)
        elif isinstance(value, Decimal):
            return str(value)
        else:
            return str(value)
    
    def _excel_serialize_value(self, value: Any) -> Any:
        """Serialize value for Excel cell."""
        if value is None:
            return None
        elif isinstance(value, (list, dict)):
            return json.dumps(value)
        elif isinstance(value, UUID):
            return str(value)
        elif isinstance(value, Decimal):
            return float(value)
        else:
            return value
    
    def _parquet_serialize_value(self, value: Any) -> Any:
        """Serialize value for Parquet."""
        if isinstance(value, UUID):
            return str(value)
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, (list, dict)):
            return json.dumps(value)
        else:
            return value
    
    def _json_serializer(self, obj: Any) -> Any:
        """JSON serializer for complex types."""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, date):
            return obj.isoformat()
        elif isinstance(obj, UUID):
            return str(obj)
        elif isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, bytes):
            return obj.decode('utf-8', errors='replace')
        elif hasattr(obj, '__dict__'):
            return obj.__dict__
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    # =========================================================================
    # FILE-BASED EXPORT METHODS (write to file path)
    # =========================================================================
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
        delimiter: str = ",",
        include_header: bool = True,
    ) -> str:
        """
        Export data to CSV file.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the CSV file
            schema: Optional schema definition
            delimiter: CSV delimiter (default comma)
            include_header: Whether to include header row
        
        Returns:
            Path to the created file
        """
        self.options.csv_delimiter = delimiter
        self.options.csv_include_header = include_header
        
        content = self.export_csv(data)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
        sheet_name: str = "Data",
    ) -> str:
        """
        Export data to Excel file.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the Excel file
            schema: Optional schema definition
            sheet_name: Name of the worksheet
        
        Returns:
            Path to the created file
        """
        self.options.excel_sheet_name = sheet_name
        
        content = self.export_excel(data, schema)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path
    
    def export_to_parquet(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
        compression: str = "snappy",
    ) -> str:
        """
        Export data to Parquet file.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the Parquet file
            schema: Optional schema definition
            compression: Compression algorithm (snappy, gzip, none)
        
        Returns:
            Path to the created file
        """
        self.options.parquet_compression = compression
        
        content = self.export_parquet(data, schema)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path
    
    def export_to_json(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
        pretty: bool = False,
        indent: Optional[int] = None,
    ) -> str:
        """
        Export data to JSON file.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the JSON file
            schema: Optional schema definition (for compatibility)
            pretty: Whether to format with indentation
            indent: Number of spaces for indentation (overrides pretty)
        
        Returns:
            Path to the created file
        """
        if indent is not None:
            self.options.json_indent = indent
        elif pretty:
            self.options.json_indent = 2
        else:
            self.options.json_indent = None
        
        content = self.export_json(data)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path
    
    def export_to_jsonl(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Export data to JSON Lines file.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the JSONL file
            schema: Optional schema definition (for compatibility)
        
        Returns:
            Path to the created file
        """
        content = self.export_jsonl(data)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path
    
    def export_to_sql(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        schema: Optional[Dict[str, Any]] = None,
        table_name: str = "data",
        dialect: str = "postgresql",
        batch_size: int = 1000,
    ) -> str:
        """
        Export data to SQL file with INSERT statements.
        
        Args:
            data: List of dictionaries containing the data
            output_path: Path to write the SQL file
            schema: Optional schema definition
            table_name: Name of the table for INSERT statements
            dialect: SQL dialect (postgresql, mysql, sqlite)
            batch_size: Number of rows per INSERT statement
        
        Returns:
            Path to the created file
        """
        self.options.sql_table_name = table_name
        self.options.sql_dialect = dialect
        self.options.sql_batch_size = batch_size
        
        content = self.export_sql(data, schema)
        
        with open(output_path, 'wb') as f:
            f.write(content)
        
        return output_path


# ============================================================================
# STREAMING EXPORT (for large datasets)
# ============================================================================

    def convert_format(
        self,
        input_path: str,
        output_path: str,
        from_format: str,
        to_format: str,
    ) -> str:
        """
        Convert a file from one format to another.
        
        Args:
            input_path: Path to the input file
            output_path: Path to write the output file
            from_format: Source format (csv, json, parquet, excel)
            to_format: Target format (csv, json, parquet, excel)
        
        Returns:
            Path to the created file
        """
        import pandas as pd
        
        # Read the source file
        if from_format == "csv":
            df = pd.read_csv(input_path)
        elif from_format == "json":
            df = pd.read_json(input_path)
        elif from_format == "parquet":
            df = pd.read_parquet(input_path)
        elif from_format in ("excel", "xlsx"):
            df = pd.read_excel(input_path)
        else:
            raise ValueError(f"Unsupported source format: {from_format}")
        
        # Convert to list of dicts
        data = df.to_dict(orient="records")
        
        # Export to target format
        if to_format == "csv":
            return self.export_to_csv(data, output_path)
        elif to_format == "json":
            return self.export_to_json(data, output_path)
        elif to_format == "parquet":
            return self.export_to_parquet(data, output_path)
        elif to_format in ("excel", "xlsx"):
            return self.export_to_excel(data, output_path)
        else:
            raise ValueError(f"Unsupported target format: {to_format}")


class StreamingExportService(ExportService):
    """Export service with streaming support for large datasets."""
    
    def export_csv_streaming(
        self,
        data_iterator: Iterator[Dict[str, Any]],
        fieldnames: List[str]
    ) -> Iterator[bytes]:
        """Stream CSV export row by row."""
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=self.options.csv_delimiter,
            quotechar=self.options.csv_quotechar,
            quoting=csv.QUOTE_MINIMAL,
        )
        
        if self.options.csv_include_header:
            writer.writeheader()
            yield output.getvalue().encode(self.options.csv_encoding)
            output.seek(0)
            output.truncate()
        
        for row in data_iterator:
            processed_row = {k: self._serialize_value(v) for k, v in row.items()}
            writer.writerow(processed_row)
            yield output.getvalue().encode(self.options.csv_encoding)
            output.seek(0)
            output.truncate()
    
    def export_jsonl_streaming(
        self,
        data_iterator: Iterator[Dict[str, Any]]
    ) -> Iterator[bytes]:
        """Stream JSONL export row by row."""
        for row in data_iterator:
            line = json.dumps(row, default=self._json_serializer, ensure_ascii=False)
            yield (line + "\n").encode('utf-8')


def get_export_service(options: Optional[ExportOptions] = None) -> ExportService:
    """Factory function to get ExportService instance."""
    return ExportService(options)
